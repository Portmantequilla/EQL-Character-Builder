#!/usr/bin/env python
"""import_stat_estimator.py — import the community HP/mana base curves from
"Mosscovered Legend's EQL Stat Estimator" (xlsx) into wiki.db.

Source model (extracted from the workbook's own formulas, v0.1.4):
  per class c at level L:  HP_c   = INT(hp[L,c]   + hp_fac[L,c]   x adjSTA)
                           Mana_c = INT(mana[L,c] + mana_fac[L,c] x convStat)
  adjSTA: STA>255 -> ROUND((STA-255)/2)+255
  convStat (INT or WIS by the class's mana type): <=100 as-is;
     101..200 -> ROUND((5*stat-300)/2); >200 -> ROUND((5*ROUND((stat+200)/2)-300)/2)
  totals: HP = 5 + top2(HP_c) summed; Mana = top2(Mana_c >= 0) summed.
  (Validated 2026-07-23 vs two live screenshots: mana within ~2-3%, HP ~4-7%.)

Writes table class_base_curve(level, class, hp, hp_fac, mana, mana_fac) — 1,600 rows
(16 classes x levels 1..100) — plus a provenance row. RE-RUN after re-syncs (the table
survives reparse, but re-run if the estimator sheet updates).

Usage: python scripts/import_stat_estimator.py [--xlsx PATH] [--db db/eql.db]
"""
import argparse
import os
import sqlite3
import sys

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DB = os.environ.get("EQL_DB", os.path.join(BASE, "db", "eql.db"))
# Pass the estimator workbook with --xlsx, or set EQL_ESTIMATOR_XLSX, or drop it in raw/.
DEFAULT_XLSX = os.environ.get(
    "EQL_ESTIMATOR_XLSX",
    os.path.join(BASE, "raw", "EQL Stat Estimator.xlsx"),
)
# the workbook's PEQ class ids -> our abbreviations
CLASS_BY_ID = {1: "WAR", 2: "CLR", 3: "PAL", 4: "RNG", 5: "SHD", 6: "DRU", 7: "MNK",
               8: "BRD", 9: "ROG", 10: "SHM", 11: "NEC", 12: "WIZ", 13: "MAG",
               14: "ENC", 15: "BST", 16: "BER"}


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", default=DEFAULT_XLSX)
    ap.add_argument("--db", default=DB)
    args = ap.parse_args()
    try:
        import openpyxl
    except ImportError:
        sys.exit("pip install openpyxl")
    if not os.path.exists(args.xlsx):
        sys.exit(f"estimator workbook not found: {args.xlsx}")

    wb = openpyxl.load_workbook(args.xlsx, data_only=True)
    bd = wb["base_data"]
    rows = []
    for r in range(2, bd.max_row + 1):
        lvl, cid, hp, hpf, mana, manaf = (bd.cell(r, c).value for c in range(1, 7))
        if lvl is None or cid is None:
            continue
        cls = CLASS_BY_ID.get(int(cid))
        if cls is None:
            continue
        rows.append((int(lvl), cls, float(hp or 0), float(hpf or 0),
                     float(mana or 0), float(manaf or 0)))
    if len(rows) < 1500:
        sys.exit(f"only {len(rows)} curve rows parsed — sheet layout changed?")

    con = sqlite3.connect(args.db)
    cur = con.cursor()
    cur.execute("DROP TABLE IF EXISTS class_base_curve")
    cur.execute(
        """CREATE TABLE class_base_curve (
             level INTEGER NOT NULL, class TEXT NOT NULL,
             hp REAL, hp_fac REAL, mana REAL, mana_fac REAL,
             PRIMARY KEY (level, class))"""
    )
    cur.executemany("INSERT INTO class_base_curve VALUES (?,?,?,?,?,?)", rows)
    cur.execute(
        """CREATE TABLE IF NOT EXISTS estimator_provenance
           (id INTEGER PRIMARY KEY CHECK (id=1), source TEXT, imported_at TEXT)"""
    )
    cur.execute(
        "INSERT OR REPLACE INTO estimator_provenance VALUES (1, ?, datetime('now'))",
        (os.path.basename(args.xlsx),),
    )
    con.commit()
    n = cur.execute("SELECT COUNT(*) FROM class_base_curve").fetchone()[0]
    lv = cur.execute("SELECT MIN(level), MAX(level) FROM class_base_curve").fetchone()
    print(f"class_base_curve: {n} rows, levels {lv[0]}..{lv[1]}, "
          f"{len(set(r[1] for r in rows))} classes")


if __name__ == "__main__":
    main()
